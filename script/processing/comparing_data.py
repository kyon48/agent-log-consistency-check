import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from pathlib import Path

ROOT_DIR = Path(__file__).parent.parent.parent

terminal_name_mapping = {
    "BCTHD010": "bct",
    "BNCTC050": "bnct",
    "BICTC010": "bptg",
    "PECTC050": "bpts",
    "HJNPC010": "hjnc",
    "HPNTC050": "hpnt",
    "PNCOC010": "pnc",
    "PNITC050": "pnit"
}

def compare_data(terminal_file, porti_file, output_file):
    # 데이터 읽기
    terminal_df = pd.read_csv(terminal_file, encoding='utf-8')
    porti_df = pd.read_csv(porti_file, encoding='utf-8')

    # 공통 컬럼 선택
    common_columns = terminal_df.columns.intersection(porti_df.columns)
    terminal_df = terminal_df[common_columns]
    porti_df = porti_df[common_columns]

    # 두 데이터프레임을 인덱스로 병합
    merged_df = terminal_df.merge(porti_df, left_index=True, right_index=True, suffixes=('_terminal', '_porti'))

    # 컬럼을 교차로 정렬
    interleaved_columns = []
    for col in common_columns:
        interleaved_columns.append(f"{col}_terminal")
        interleaved_columns.append(f"{col}_porti")
    merged_df = merged_df[interleaved_columns]

    # # etb_terminal 기준으로 오름차순 정렬
    # if 'etb_terminal' in merged_df.columns:
    #     merged_df = merged_df.sort_values(by='etb_terminal')

    # 엑셀 파일 생성
    wb = Workbook()
    ws = wb.active

    # 헤더 추가
    ws.append(['Index'] + list(merged_df.columns))

    # 색상 설정
    green_fill = PatternFill(start_color="BBFFBB", end_color="BBFFBB", fill_type="solid")
    red_fill = PatternFill(start_color="FFBBBB", end_color="FFBBBB", fill_type="solid")

    # 데이터 추가 및 색상 적용
    for row_num, (index, row) in enumerate(merged_df.iterrows(), 2):
        ws.append([index] + list(row))
        for col_num, col_name in enumerate(merged_df.columns, 2):
            cell = ws.cell(row=row_num, column=col_num)
            # 컬럼 이름에서 접미사를 제거하고 비교
            base_col_name = col_name.rsplit('_', 1)[0]
            terminal_value = terminal_df.at[index, base_col_name]
            porti_value = porti_df.at[index, base_col_name]
            # 두 값이 같거나 둘 다 비어 있는 경우 초록색
            if terminal_value == porti_value or (pd.isna(terminal_value) and pd.isna(porti_value)):
                cell.fill = green_fill
            else:
                cell.fill = red_fill

    # 엑셀 파일 저장
    wb.save(output_file)
    print(f"정합성 비교 결과가 {output_file}에 저장되었습니다.")

if __name__ == "__main__":
    for _, terminal_name in terminal_name_mapping.items():
        terminal_file = ROOT_DIR / "processed_data" / f"processed_{terminal_name}.csv"
        porti_file = ROOT_DIR / "processed_data" / f"processed_porti_{terminal_name}.csv"
        output_file = ROOT_DIR / "comparison_results" / f"comparison_{terminal_name}.xlsx"
        compare_data(terminal_file, porti_file, output_file)