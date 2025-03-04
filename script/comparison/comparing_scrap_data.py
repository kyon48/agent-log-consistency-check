import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from pathlib import Path

ROOT_DIR = Path(__file__).parent.parent.parent

terminal_name_mapping = {
    "BCTHD010": "bct",
    "BNCTC050": "bnct",
    "BICTC010": "bptg",
    "PECTC050": "bpts",
    "HJNPC010": "hjnc",
    "HPNTC050": "hpnt",
    "PNCOC010": "pnc",
    "PNITC050": "pnit",
    "GCTOC050": "hktg",
    "DGTBC050": "dgt"
}

def compare_data(scrap_file, porti_file, output_file):
    # 데이터 읽기
    scrap_df = pd.read_csv(scrap_file, encoding='utf-8')
    porti_df = pd.read_csv(porti_file, encoding='utf-8')

    # terminalShipVoyageNo를 기준으로 병합
    merged_df = scrap_df.merge(porti_df, on='terminalShipVoyageNo', how='outer', suffixes=('_scrap', '_porti'))

    # terminalShipVoyageNo를 맨 앞에 두고 나머지 컬럼을 사전 순으로 정렬
    sorted_columns = ['terminalShipVoyageNo'] + sorted([col for col in merged_df.columns if col != 'terminalShipVoyageNo'])

    # 엑셀 파일 생성
    wb = Workbook()
    ws = wb.active

    # 헤더 추가
    ws.append(sorted_columns)

    # 색상 설정
    green_fill = PatternFill(start_color="BBFFBB", end_color="BBFFBB", fill_type="solid")
    red_fill = PatternFill(start_color="FFBBBB", end_color="FFBBBB", fill_type="solid")

    # 데이터 추가 및 색상 적용
    for row_num, row in enumerate(merged_df[sorted_columns].itertuples(index=False), 2):
        ws.append(row)
        for col_num, col_name in enumerate(sorted_columns[1:], 2):  # 첫 번째 컬럼은 terminalShipVoyageNo이므로 제외
            cell = ws.cell(row=row_num, column=col_num)
            # 컬럼 이름에서 접미사를 제거하고 비교
            base_col_name = col_name.rsplit('_', 1)[0]
            scrap_value = getattr(row, f"{base_col_name}_scrap", None)
            porti_value = getattr(row, f"{base_col_name}_porti", None)
            # 두 값이 같거나 둘 다 비어 있는 경우 초록색
            if scrap_value == porti_value or (pd.isna(scrap_value) and pd.isna(porti_value)):
                cell.fill = green_fill
            else:
                cell.fill = red_fill
    # 엑셀 파일 저장
    wb.save(output_file)
    print(f"정합성 비교 결과가 {output_file}에 저장되었습니다.")


if __name__ == "__main__":
    scrap_file = ROOT_DIR / "processed_data" / "scrapdb" / f"processed_scrapdb.csv"
    porti_file = ROOT_DIR / "processed_data" / "porti" / f"processed_porti.csv"
    output_file = ROOT_DIR / "comparison_results" / f"comparison_scrapdb_porti.xlsx"
    compare_data(scrap_file, porti_file, output_file)

    for _, terminal_name in terminal_name_mapping.items():
        scrap_file = ROOT_DIR / "processed_data" / "scrapdb" / f"processed_scrapdb_{terminal_name}.csv"
        porti_file = ROOT_DIR / "processed_data" / "porti" / f"processed_porti_{terminal_name}.csv"
        output_file = ROOT_DIR / "comparison_results" / f"comparison_scrapdb_porti_{terminal_name}.xlsx"
        compare_data(scrap_file, porti_file, output_file)